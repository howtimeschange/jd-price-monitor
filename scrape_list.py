"""
抓取亚瑟士京东店铺商品列表，输出 Excel
用 bb-browser adapter 逐页抓取（需要 Chrome 已登录 JD，bb-browser daemon 运行中）
"""
import json
import os
import re
import subprocess
import time

SHOP_ID = "1000462158"
VENDOR_ID = "2863474"
PAGE_SIZE = 60
CDP_PORT = "9222"
BASE_URL = "https://mall.jd.com/advance_search-{vendor_id}-{shop_id}-{shop_id}-0-0-0-1-{page}-{page_size}.html"
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "data", "asics_price_list.xlsx")


def bb(args, timeout=15):
    """运行 bb-browser 命令"""
    return subprocess.run(["bb-browser"] + args + ["--port", CDP_PORT],
                          capture_output=True, text=True, timeout=timeout)


def get_jd_tab() -> str:
    """动态找到 mall.jd.com 的 tab index"""
    r = bb(["tab", "list"])
    for line in r.stdout.splitlines():
        if "mall.jd.com" in line:
            m = re.search(r'\[(\d+)\]', line)
            if m:
                return m.group(1)
    return "0"


def navigate_and_wait(url: str) -> bool:
    """切换到 JD tab，导航到 URL，等待商品列表加载"""
    tab = get_jd_tab()
    # 切换激活 tab（site 命令会在激活 tab 执行）
    bb(["tab", tab])
    # 导航
    bb(["eval", f"location.href='{url}'", "--tab", tab])
    # 轮询等待 jSubObject 出现（最多 25 秒）
    for _ in range(25):
        time.sleep(1)
        r = bb(["eval", "document.querySelectorAll('li.jSubObject').length", "--tab", tab], timeout=5)
        try:
            if int(r.stdout.strip()) > 0:
                time.sleep(5)  # 等价格 JS 渲染（保守 5 秒）
                return True
        except Exception:
            pass
    # 超时后额外等 5 秒
    time.sleep(5)
    return False


def scrape_current_page() -> dict:
    """用 bb-browser adapter 抓取当前页 DOM（含滚动触发价格渲染）"""
    # site 命令不支持 --tab，直接不带
    r = bb(["site", "jd/shop-prices", "--json"], timeout=30)
    if r.returncode != 0:
        return {"error": r.stderr.strip(), "items": []}
    try:
        out = r.stdout.strip()
        start = out.find('{')
        if start > 0:
            out = out[start:]
        parsed = json.loads(out)
        if not parsed.get("success", True):
            return {"error": parsed.get("error", "unknown"), "items": []}
        if "data" in parsed:
            return parsed["data"]
        return parsed
    except Exception as e:
        return {"error": str(e), "raw": r.stdout[:200], "items": []}


def main():
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    all_items = {}
    page_no = 1
    consecutive_empty = 0

    print(f"开始抓取亚瑟士京东店铺（shop_id={SHOP_ID}）...")
    print(f"请确认：Chrome 已打开并登录 JD，bb-browser daemon 运行中（port {CDP_PORT}）\n")

    # 导航到第一页
    start_url = BASE_URL.format(vendor_id=VENDOR_ID, shop_id=SHOP_ID, page=1, page_size=PAGE_SIZE)
    print(f"  导航到第 1 页...", end=" ", flush=True)
    print("已加载" if navigate_and_wait(start_url) else "超时，继续尝试")

    while True:
        print(f"  抓取第 {page_no} 页数据...", end=" ", flush=True)
        data = scrape_current_page()

        if "error" in data and not data.get("items"):
            print(f"出错: {data['error']}")
            consecutive_empty += 1
            if consecutive_empty >= 2:
                print("  连续两页失败，停止")
                break
            page_no += 1
            continue

        items = data.get("items", [])
        missing = data.get("missingCount", 0)
        print(f"{len(items)} 个商品，有价格: {data.get('withPrice', len(items) - missing)}，缺失: {missing}")

        if not items:
            consecutive_empty += 1
            if consecutive_empty >= 2:
                print("  连续两页无商品，停止")
                break
        else:
            consecutive_empty = 0
            for item in items:
                if item["skuId"] not in all_items:
                    all_items[item["skuId"]] = item

        next_url = data.get("nextUrl")
        if not next_url:
            print("  已到最后一页")
            break

        page_no += 1
        print(f"  导航到第 {page_no} 页...", end=" ", flush=True)
        print("已加载" if navigate_and_wait(next_url) else "超时，继续等待...")
        time.sleep(2)  # 超时后额外等待

    print(f"\n共抓取 {len(all_items)} 个商品，正在写入 Excel...")

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        subprocess.run(["pip3", "install", "openpyxl", "-q"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "价格列表"

    headers = ["款号(SKU ID)", "商品名称", "页面价(元)", "原价(元)", "商品链接"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(all_items.values(), 2):
        ws.cell(row=row_idx, column=1, value=item["skuId"])
        ws.cell(row=row_idx, column=2, value=item.get("name", ""))
        try:
            ws.cell(row=row_idx, column=3, value=float(item["price"]) if item.get("price") else "")
        except Exception:
            ws.cell(row=row_idx, column=3, value=item.get("price", ""))
        try:
            ws.cell(row=row_idx, column=4, value=float(item["originalPrice"]) if item.get("originalPrice") else "")
        except Exception:
            ws.cell(row=row_idx, column=4, value=item.get("originalPrice", ""))
        ws.cell(row=row_idx, column=5, value=item.get("href", ""))
        if row_idx % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="EBF3FB")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    print(f"✅ Excel 已保存：{OUTPUT_FILE}")
    print(f"   共 {len(all_items)} 条记录")


if __name__ == "__main__":
    main()
