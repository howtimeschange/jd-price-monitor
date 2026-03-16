#!/usr/bin/env python3
"""
JD Price Monitor — 交互式命令行界面
运行方式：python cli.py
"""

import os
import re
import sys
import time
import asyncio
import subprocess
import threading
from datetime import datetime, timedelta
from pathlib import Path

# ── 依赖检查 ────────────────────────────────────────────────────────────────
def _ensure_deps():
    required = {"rich": "rich", "questionary": "questionary", "openpyxl": "openpyxl"}
    missing = []
    for mod, pkg in required.items():
        try:
            __import__(mod)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"正在安装依赖: {', '.join(missing)} ...")
        subprocess.run([sys.executable, "-m", "pip", "install", "-q"] + missing, check=True)

_ensure_deps()

import questionary
from questionary import Style
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.text import Text
from rich.live import Live
from rich.progress import Progress, SpinnerColumn, TextColumn, BarColumn, TaskProgressColumn, TimeElapsedColumn
from rich.prompt import Prompt, Confirm
from rich import box
from rich.columns import Columns
from rich.padding import Padding
from rich.rule import Rule

sys.path.insert(0, os.path.dirname(__file__))
from src.config import load_config, save_config, reload_config

console = Console()

PROJ_DIR = Path(__file__).parent

# ── 问答样式 ────────────────────────────────────────────────────────────────
Q_STYLE = Style([
    ("qmark",         "fg:#4FC3F7 bold"),
    ("question",      "bold"),
    ("answer",        "fg:#80CBC4 bold"),
    ("pointer",       "fg:#4FC3F7 bold"),
    ("highlighted",   "fg:#4FC3F7 bold"),
    ("selected",      "fg:#80CBC4"),
    ("separator",     "fg:#6C6C6C"),
    ("instruction",   "fg:#6C6C6C italic"),
])

# ── 工具函数 ────────────────────────────────────────────────────────────────

def clear():
    os.system("cls" if os.name == "nt" else "clear")


def print_banner():
    banner = Text()
    banner.append("  JD Price Monitor", style="bold cyan")
    banner.append("  京东价格监控系统", style="dim")
    console.print(Panel(banner, border_style="cyan", padding=(0, 2)))
    console.print()


def print_current_config():
    """在菜单上方显示当前配置摘要"""
    try:
        cfg = reload_config()
        shop = cfg.get("shop", {})
        mon  = cfg.get("monitor", {})
        dt   = cfg.get("dingtalk", {})

        t = Table(box=box.SIMPLE, show_header=False, padding=(0, 1))
        t.add_column("key",   style="dim", width=12)
        t.add_column("value", style="cyan")

        shop_url = shop.get("shop_url", "")
        # 从 URL 提取 shop_id 显示
        m = re.search(r"index-(\d+)", shop_url)
        shop_display = m.group(1) if m else shop_url[:40] or "未配置"

        t.add_row("店铺",     shop.get("shop_name", shop_display))
        t.add_row("阈值",     f"{int(mon.get('price_ratio_threshold', 0.5) * 100)}折  ({mon.get('price_ratio_threshold', 0.5):.0%})")
        t.add_row("巡检间隔", f"{mon.get('interval_minutes', 120)} 分钟")
        webhook = dt.get("webhook_url", "")
        t.add_row("Webhook", "✅ 已配置" if webhook and "YOUR_TOKEN" not in webhook else "⚠️  未配置")

        console.print(Panel(t, title="[dim]当前配置[/dim]", border_style="dim", padding=(0, 1)))
        console.print()
    except Exception:
        pass


MENU_CHOICES = [
    {"name": "📦  导出全店商品价格  →  Excel", "value": "export"},
    {"name": "🔍  立即执行一次破价巡检",        "value": "check_once"},
    {"name": "🔁  循环巡检（按间隔自动运行）",  "value": "loop"},
    {"name": "⏰  创建系统定时任务（cron）",    "value": "cron"},
    {"name": "⚙️   设置  —  店铺 / 阈值 / Webhook", "value": "settings"},
    {"name": "──────────────────────────────",  "value": "sep", "disabled": ""},
    {"name": "❌  退出",                         "value": "exit"},
]


# ═══════════════════════════════════════════════════════════════════
# ① 导出 Excel
# ═══════════════════════════════════════════════════════════════════

def _run_export_with_progress():
    """在后台线程跑 scrape_list 核心逻辑，同时显示 rich 进度条"""
    import json, re as _re, subprocess as _sp

    cfg  = reload_config()
    shop = cfg.get("shop", {})

    # 从 shop_url 解析 shop_id / vendor_id
    shop_url = shop.get("shop_url", "")
    m_shop   = _re.search(r"index-(\d+)", shop_url)
    shop_id  = m_shop.group(1) if m_shop else cfg["shop"].get("shop_id", "")

    # vendor_id 从 advance_search URL 里取不到，用 config 兜底
    vendor_id = cfg["shop"].get("vendor_id", shop_id)

    cdp_port  = str(cfg.get("cdp_port", 9222))
    page_size = 60
    base_url  = (
        f"https://mall.jd.com/advance_search-{vendor_id}-{shop_id}"
        f"-{shop_id}-0-0-0-1-{{page}}-{page_size}.html"
    )
    out_dir  = PROJ_DIR / cfg["output"].get("data_dir", "data")
    out_dir.mkdir(parents=True, exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = out_dir / f"price_list_{ts}.xlsx"

    def bb(args, timeout=15):
        return _sp.run(
            ["bb-browser"] + args + ["--port", cdp_port],
            capture_output=True, text=True, timeout=timeout,
        )

    def get_jd_tab():
        r = bb(["tab", "list"])
        for line in r.stdout.splitlines():
            if "mall.jd.com" in line:
                mx = _re.search(r'\[(\d+)\]', line)
                if mx:
                    return mx.group(1)
        return "0"

    def navigate_and_wait(url):
        import time as _t
        tab = get_jd_tab()
        bb(["tab", tab])
        bb(["eval", f"location.href='{url}'", "--tab", tab])
        for _ in range(25):
            _t.sleep(1)
            r = bb(["eval", "document.querySelectorAll('li.jSubObject').length", "--tab", tab], timeout=5)
            try:
                if int(r.stdout.strip()) > 0:
                    _t.sleep(5)   # 保守等待 5 秒让价格渲染
                    return True
            except Exception:
                pass
        _t.sleep(5)
        return False

    def scrape_page():
        r = bb(["site", "jd/shop-prices", "--json"], timeout=60)
        if r.returncode != 0:
            return {"error": r.stderr.strip(), "items": []}
        try:
            out = r.stdout.strip()
            s   = out.find("{")
            if s > 0:
                out = out[s:]
            parsed = json.loads(out)
            if not parsed.get("success", True):
                return {"error": parsed.get("error", "unknown"), "items": []}
            return parsed.get("data", parsed)
        except Exception as e:
            return {"error": str(e), "items": []}

    # ── 开始抓取 ──
    all_items = {}
    page_no   = 1
    messages  = []

    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        BarColumn(bar_width=30),
        TaskProgressColumn(),
        TimeElapsedColumn(),
        console=console,
        transient=False,
    ) as progress:
        task = progress.add_task("[cyan]正在抓取...", total=None)

        def upd(msg):
            progress.update(task, description=f"[cyan]{msg}")
            messages.append(msg)

        upd(f"导航到第 1 页...")
        start_url = base_url.format(page=1)
        ok = navigate_and_wait(start_url)
        upd(f"第 1 页{'已加载' if ok else '超时，继续'}")

        while True:
            upd(f"抓取第 {page_no} 页...")
            data  = scrape_page()
            items = data.get("items", [])

            if items:
                for item in items:
                    if item["skuId"] not in all_items:
                        all_items[item["skuId"]] = item
                missing = data.get("missingCount", 0)
                upd(
                    f"第 {page_no} 页 ✓  {len(items)} 个商品，"
                    f"有价格 {len(items)-missing}，缺失 {missing}  "
                    f"（累计 {len(all_items)}）"
                )
            else:
                upd(f"第 {page_no} 页无商品数据，停止")
                break

            next_url = data.get("nextUrl")
            if not next_url:
                upd("已到最后一页 ✅")
                break

            page_no += 1
            upd(f"导航到第 {page_no} 页...")
            ok = navigate_and_wait(next_url)
            upd(f"第 {page_no} 页{'已加载' if ok else '超时，继续'}")

        progress.update(task, description=f"[green]完成，共 {len(all_items)} 个商品，正在写 Excel...")

    # ── 写 Excel ──
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        _sp.run([sys.executable, "-m", "pip", "install", "-q", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "价格列表"

    headers     = ["款号(SKU ID)", "商品名称", "页面价(元)", "原价(元)", "商品链接"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, item in enumerate(all_items.values(), 2):
        ws.cell(row=row_idx, column=1, value=item["skuId"])
        ws.cell(row=row_idx, column=2, value=item.get("name", ""))
        try:
            ws.cell(row=row_idx, column=3, value=float(item["price"]) if item.get("price") else "")
        except Exception:
            ws.cell(row=row_idx, column=3, value=item.get("price", ""))
        try:
            ws.cell(row=row_idx, column=4, value=float(item["originalPrice"]) if item.get("originalPrice") else "")
        except Exception:
            ws.cell(row=row_idx, column=4, value=item.get("originalPrice", ""))
        ws.cell(row=row_idx, column=5, value=item.get("href", ""))
        if row_idx % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="EBF3FB")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 55
    ws.freeze_panes = "A2"
    wb.save(str(out_file))

    return len(all_items), str(out_file)


def action_export():
    console.print(Rule("[cyan]导出全店商品价格[/cyan]"))
    console.print(
        "[dim]将通过 bb-browser 逐页抓取商品列表，导出到 Excel。\n"
        "请确保：Chrome 已登录 JD，bb-browser daemon 运行中（port 9222）[/dim]\n"
    )

    if not questionary.confirm("确认开始抓取？", style=Q_STYLE, default=True).ask():
        return

    try:
        count, out_file = _run_export_with_progress()
        console.print()
        console.print(
            Panel(
                f"[green bold]✅ 导出完成[/green bold]\n\n"
                f"  共 [cyan bold]{count}[/cyan bold] 个商品\n"
                f"  文件：[link=file://{out_file}]{out_file}[/link]",
                border_style="green",
                padding=(1, 2),
            )
        )
    except Exception as e:
        console.print(f"\n[red]❌ 导出失败：{e}[/red]")

    console.print()
    Prompt.ask("[dim]按 Enter 返回主菜单[/dim]")


# ═══════════════════════════════════════════════════════════════════
# ② 单次破价巡检
# ═══════════════════════════════════════════════════════════════════

def action_check_once():
    console.print(Rule("[cyan]立即执行一次破价巡检[/cyan]"))
    console.print("[dim]使用 main.py 的巡检逻辑，完成后显示结果[/dim]\n")

    if not questionary.confirm("确认开始巡检？", style=Q_STYLE, default=True).ask():
        return

    console.print()
    with Progress(
        SpinnerColumn(),
        TextColumn("[progress.description]{task.description}"),
        TimeElapsedColumn(),
        console=console,
        transient=True,
    ) as progress:
        task = progress.add_task("[cyan]正在巡检...", total=None)

        # 在子进程跑，避免事件循环冲突
        result = subprocess.run(
            [sys.executable, str(PROJ_DIR / "main.py")],
            capture_output=True, text=True,
            cwd=str(PROJ_DIR),
        )
        progress.update(task, description="[green]巡检完成")

    if result.returncode == 0:
        console.print(Panel(
            f"[green bold]✅ 巡检完成[/green bold]\n\n[dim]{result.stdout[-1500:]}[/dim]",
            border_style="green", padding=(1, 2),
        ))
    else:
        console.print(Panel(
            f"[red bold]❌ 巡检出错[/red bold]\n\n[dim]{result.stderr[-1000:]}[/dim]",
            border_style="red", padding=(1, 2),
        ))

    console.print()
    Prompt.ask("[dim]按 Enter 返回主菜单[/dim]")


# ═══════════════════════════════════════════════════════════════════
# ③ 循环巡检
# ═══════════════════════════════════════════════════════════════════

def action_loop():
    cfg      = reload_config()
    interval = cfg["monitor"]["interval_minutes"]
    console.print(Rule("[cyan]循环巡检[/cyan]"))
    console.print(
        f"[dim]将每 [cyan bold]{interval}[/cyan bold] 分钟执行一次巡检。\n"
        "按 [bold]Ctrl+C[/bold] 停止。[/dim]\n"
    )

    run_in = questionary.select(
        "请选择运行方式：",
        choices=[
            {"name": "在当前终端前台运行（Ctrl+C 停止）",          "value": "fg"},
            {"name": "以后台进程运行（输出到 logs/loop.log）",     "value": "bg"},
            {"name": "返回",                                        "value": "back"},
        ],
        style=Q_STYLE,
    ).ask()

    if run_in == "back" or run_in is None:
        return

    if run_in == "fg":
        console.print("\n[dim]启动前台循环巡检，Ctrl+C 停止...[/dim]\n")
        try:
            subprocess.run(
                [sys.executable, str(PROJ_DIR / "main.py"), "--loop"],
                cwd=str(PROJ_DIR),
            )
        except KeyboardInterrupt:
            console.print("\n[yellow]已停止循环巡检[/yellow]")
    else:
        log_dir  = PROJ_DIR / cfg["output"].get("log_dir", "logs")
        log_dir.mkdir(parents=True, exist_ok=True)
        log_file = log_dir / "loop.log"
        proc = subprocess.Popen(
            [sys.executable, str(PROJ_DIR / "main.py"), "--loop"],
            stdout=open(str(log_file), "a"),
            stderr=subprocess.STDOUT,
            cwd=str(PROJ_DIR),
            start_new_session=True,
        )
        console.print(Panel(
            f"[green bold]✅ 后台进程已启动[/green bold]\n\n"
            f"  PID：[cyan]{proc.pid}[/cyan]\n"
            f"  日志：[link=file://{log_file}]{log_file}[/link]\n\n"
            f"  停止命令：[bold]kill {proc.pid}[/bold]",
            border_style="green", padding=(1, 2),
        ))
        Prompt.ask("\n[dim]按 Enter 返回主菜单[/dim]")


# ═══════════════════════════════════════════════════════════════════
# ④ 定时任务（cron）
# ═══════════════════════════════════════════════════════════════════

def action_cron():
    cfg      = reload_config()
    interval = cfg["monitor"]["interval_minutes"]
    console.print(Rule("[cyan]创建系统定时任务[/cyan]"))

    # 计算 cron 表达式
    if interval >= 60 and 60 % (interval // 60) == 0:
        every_h = interval // 60
        cron_expr = f"0 */{every_h} * * *"
        cron_desc = f"每 {every_h} 小时"
    else:
        cron_expr = f"*/{interval} * * * *"
        cron_desc = f"每 {interval} 分钟"

    python_bin = sys.executable
    proj_path  = str(PROJ_DIR)
    log_dir    = PROJ_DIR / cfg["output"].get("log_dir", "logs")
    log_dir.mkdir(parents=True, exist_ok=True)
    log_file   = log_dir / "cron.log"

    cron_line = (
        f"{cron_expr}  cd \"{proj_path}\" && "
        f"\"{python_bin}\" main.py >> \"{log_file}\" 2>&1"
    )

    console.print(f"\n建议的 crontab 条目（[dim]{cron_desc}[/dim]）：\n")
    console.print(Panel(cron_line, border_style="cyan", padding=(0, 1)))

    action = questionary.select(
        "操作：",
        choices=[
            {"name": "自动写入 crontab（追加）",   "value": "add"},
            {"name": "仅复制到剪贴板",              "value": "copy"},
            {"name": "返回",                        "value": "back"},
        ],
        style=Q_STYLE,
    ).ask()

    if action == "add":
        try:
            # 读取现有 crontab
            existing = subprocess.run(["crontab", "-l"], capture_output=True, text=True)
            existing_lines = existing.stdout if existing.returncode == 0 else ""
            # 检查是否已存在
            if "main.py" in existing_lines and proj_path in existing_lines:
                console.print("[yellow]⚠️  crontab 中已存在本项目的任务，跳过添加[/yellow]")
            else:
                new_crontab = existing_lines.rstrip("\n") + "\n" + cron_line + "\n"
                proc = subprocess.run(["crontab", "-"], input=new_crontab, capture_output=True, text=True)
                if proc.returncode == 0:
                    console.print("[green]✅ 已成功写入 crontab[/green]")
                else:
                    console.print(f"[red]❌ 写入失败：{proc.stderr}[/red]")
        except FileNotFoundError:
            console.print("[red]❌ 系统未找到 crontab 命令（仅支持 macOS/Linux）[/red]")

    elif action == "copy":
        try:
            subprocess.run(["pbcopy"], input=cron_line, text=True)
            console.print("[green]✅ 已复制到剪贴板[/green]")
        except FileNotFoundError:
            try:
                subprocess.run(["xclip", "-selection", "clipboard"], input=cron_line, text=True)
                console.print("[green]✅ 已复制到剪贴板[/green]")
            except FileNotFoundError:
                console.print(f"[yellow]请手动复制上方命令[/yellow]")

    console.print()
    Prompt.ask("[dim]按 Enter 返回主菜单[/dim]")


# ═══════════════════════════════════════════════════════════════════
# ⑤ 设置
# ═══════════════════════════════════════════════════════════════════

def _parse_shop_id_from_url(url: str):
    """从京东店铺 URL 提取 shop_id，兼容多种格式"""
    patterns = [
        r"index-(\d+)\.html",
        r"shop_id=(\d+)",
        r"/(\d{9,12})(?:[/?]|$)",
    ]
    for pat in patterns:
        m = re.search(pat, url)
        if m:
            return m.group(1)
    return None


def settings_shop():
    """配置店铺 URL"""
    cfg = reload_config()
    console.print(Rule("[cyan]设置店铺[/cyan]"))
    console.print(
        "[dim]请输入京东店铺首页地址，例如：\n"
        "  https://mall.jd.com/index-1000462158.html\n"
        "  https://shop.m.jd.com/?shopId=1000462158[/dim]\n"
    )

    current_url = cfg["shop"].get("shop_url", "")
    new_url = questionary.text(
        "店铺首页 URL：",
        default=current_url,
        style=Q_STYLE,
        validate=lambda v: True if v.strip() else "不能为空",
    ).ask()
    if new_url is None:
        return

    new_url  = new_url.strip()
    shop_id  = _parse_shop_id_from_url(new_url)

    if shop_id:
        console.print(f"  [dim]已解析 shop_id：[cyan]{shop_id}[/cyan][/dim]")
    else:
        console.print("  [yellow]⚠️  无法从 URL 中解析 shop_id，请手动输入[/yellow]")
        shop_id = questionary.text("shop_id：", style=Q_STYLE).ask() or ""

    shop_name = questionary.text(
        "店铺显示名称（用于告警消息）：",
        default=cfg["shop"].get("shop_name", ""),
        style=Q_STYLE,
    ).ask() or cfg["shop"].get("shop_name", "")

    cfg["shop"]["shop_url"]  = new_url
    cfg["shop"]["shop_id"]   = shop_id
    cfg["shop"]["shop_name"] = shop_name
    save_config(cfg)
    console.print("[green]✅ 店铺配置已保存[/green]")


def settings_threshold():
    """配置破价阈值"""
    cfg = reload_config()
    console.print(Rule("[cyan]设置破价阈值[/cyan]"))
    current = cfg["monitor"].get("price_ratio_threshold", 0.5)
    console.print(f"  当前阈值：[cyan]{int(current * 100)}折[/cyan]（前台价 / 吊牌价 < {current:.0%} 则告警）\n")

    raw = questionary.text(
        "新阈值（输入折扣，例如 50 代表 5折，即 50%）：",
        default=str(int(current * 100)),
        style=Q_STYLE,
        validate=lambda v: True if v.strip().isdigit() and 1 <= int(v) <= 99 else "请输入 1-99 的整数",
    ).ask()
    if raw is None:
        return

    new_threshold = int(raw) / 100
    cfg["monitor"]["price_ratio_threshold"] = new_threshold
    save_config(cfg)
    console.print(f"[green]✅ 阈值已更新为 {int(new_threshold * 100)}折 ({new_threshold:.0%})[/green]")


def settings_webhook():
    """配置钉钉 Webhook"""
    cfg = reload_config()
    console.print(Rule("[cyan]设置钉钉 Webhook[/cyan]"))
    console.print(
        "[dim]钉钉群 → 群设置 → 智能群助手 → 添加机器人 → 自定义机器人\n"
        "复制 Webhook 地址粘贴到此处[/dim]\n"
    )

    current = cfg["dingtalk"].get("webhook_url", "")
    if "YOUR_TOKEN" in current:
        current = ""

    new_url = questionary.text(
        "Webhook URL：",
        default=current,
        style=Q_STYLE,
        validate=lambda v: True if v.strip().startswith("https://") else "请输入完整 https URL",
    ).ask()
    if new_url is None:
        return

    cfg["dingtalk"]["webhook_url"] = new_url.strip()

    # 可选：加签密钥
    use_secret = questionary.confirm("是否配置加签密钥（机器人开启了安全设置→加签）？", default=False, style=Q_STYLE).ask()
    if use_secret:
        secret = questionary.text("加签密钥（SEC...）：", style=Q_STYLE).ask() or ""
        cfg["dingtalk"]["secret"] = secret

    # 可选：@手机号
    mobiles_raw = questionary.text(
        "告警时 @的手机号（多个用逗号分隔，不需要则留空）：",
        default=",".join(cfg["dingtalk"].get("at_mobiles", [])),
        style=Q_STYLE,
    ).ask() or ""
    cfg["dingtalk"]["at_mobiles"] = [m.strip() for m in mobiles_raw.split(",") if m.strip()]

    save_config(cfg)
    console.print("[green]✅ Webhook 已保存[/green]")

    # 发一条测试消息
    if questionary.confirm("是否发送测试消息验证配置？", default=True, style=Q_STYLE).ask():
        _test_webhook(cfg)


def _test_webhook(cfg):
    import json, urllib.request, urllib.parse, hmac, hashlib, base64, time as _t
    webhook = cfg["dingtalk"]["webhook_url"]
    secret  = cfg["dingtalk"].get("secret", "")
    url     = webhook
    if secret:
        ts   = int(_t.time() * 1000)
        sign_str = f"{ts}\n{secret}"
        hmac_code = hmac.new(secret.encode(), sign_str.encode(), digestmod=hashlib.sha256).digest()
        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))
        url  = f"{webhook}&timestamp={ts}&sign={sign}"
    payload = {
        "msgtype": "markdown",
        "markdown": {
            "title": "JD Price Monitor 测试",
            "text": "## ✅ JD Price Monitor\n\n钉钉 Webhook 配置成功！监控系统已就绪。",
        },
        "at": {"isAtAll": False},
    }
    try:
        data = json.dumps(payload).encode()
        req  = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"}, method="POST")
        with urllib.request.urlopen(req, timeout=10) as resp:
            result = json.loads(resp.read().decode())
            if result.get("errcode") == 0:
                console.print("[green]✅ 测试消息发送成功[/green]")
            else:
                console.print(f"[red]❌ 钉钉返回错误：{result}[/red]")
    except Exception as e:
        console.print(f"[red]❌ 发送失败：{e}[/red]")


def settings_interval():
    """配置巡检间隔"""
    cfg = reload_config()
    console.print(Rule("[cyan]设置巡检间隔[/cyan]"))
    current = cfg["monitor"].get("interval_minutes", 120)

    choices = [
        {"name": "30 分钟", "value": 30},
        {"name": "60 分钟（1小时）", "value": 60},
        {"name": "120 分钟（2小时，推荐）", "value": 120},
        {"name": "240 分钟（4小时）", "value": 240},
        {"name": "自定义", "value": -1},
    ]
    val = questionary.select("巡检间隔：", choices=choices, style=Q_STYLE).ask()
    if val is None:
        return
    if val == -1:
        raw = questionary.text(
            "自定义分钟数（整数）：",
            validate=lambda v: True if v.strip().isdigit() and int(v) >= 5 else "请输入 ≥5 的整数",
            style=Q_STYLE,
        ).ask()
        if raw is None:
            return
        val = int(raw)

    cfg["monitor"]["interval_minutes"] = val
    save_config(cfg)
    console.print(f"[green]✅ 巡检间隔已更新为 {val} 分钟[/green]")


def action_settings():
    while True:
        console.print(Rule("[cyan]设置[/cyan]"))
        choice = questionary.select(
            "请选择要修改的配置项：",
            choices=[
                {"name": "🏪  店铺 URL / 店铺名称",   "value": "shop"},
                {"name": "📊  破价阈值",               "value": "threshold"},
                {"name": "🔔  钉钉 Webhook",           "value": "webhook"},
                {"name": "⏱  巡检间隔",               "value": "interval"},
                {"name": "↩  返回主菜单",              "value": "back"},
            ],
            style=Q_STYLE,
        ).ask()

        if choice == "shop":
            settings_shop()
        elif choice == "threshold":
            settings_threshold()
        elif choice == "webhook":
            settings_webhook()
        elif choice == "interval":
            settings_interval()
        else:
            break
        console.print()


# ═══════════════════════════════════════════════════════════════════
# 主循环
# ═══════════════════════════════════════════════════════════════════

def main():
    while True:
        clear()
        print_banner()
        print_current_config()

        choice = questionary.select(
            "请选择操作：",
            choices=MENU_CHOICES,
            style=Q_STYLE,
        ).ask()

        console.print()

        if choice == "export":
            action_export()
        elif choice == "check_once":
            action_check_once()
        elif choice == "loop":
            action_loop()
        elif choice == "cron":
            action_cron()
        elif choice == "settings":
            action_settings()
        elif choice == "exit" or choice is None:
            console.print("[dim]再见！[/dim]")
            break


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        console.print("\n[dim]已退出[/dim]")
